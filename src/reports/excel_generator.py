"""Excel report generator with charts for BookSpring metrics."""
import pandas as pd
from datetime import date
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.data.processor import DataProcessor, TimeUnit, get_friendly_name


class ExcelReportGenerator:
    """Generate Excel reports with charts from BookSpring data."""

    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self, processor: DataProcessor):
        self.processor = processor
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def _style_header_row(self, ws, row_num: int = 1):
        """Apply styling to header row."""
        for cell in ws[row_num]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _add_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int = 1):
        """Add a DataFrame to a worksheet with styling."""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.BORDER
                if r_idx == start_row:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        col_letter = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def add_time_series_sheet(self, time_unit: TimeUnit, metrics: list,
                              sheet_name: Optional[str] = None):
        """Add a sheet with time series data and line charts."""
        df = self.processor.aggregate_by_time(time_unit, metrics)
        if df.empty:
            return

        sheet_name = sheet_name or f"By {time_unit.title()}"
        ws = self.wb.create_sheet(sheet_name)

        # Rename columns to friendly names
        df = df.rename(columns={col: get_friendly_name(col) for col in df.columns})

        # Add title
        ws.cell(row=1, column=1, value=f"BookSpring Metrics by {time_unit.title()}")
        ws.cell(row=1, column=1).font = self.TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Add data starting at row 3
        self._add_dataframe_to_sheet(ws, df, start_row=3)

        # Add line chart for each metric
        chart_row = len(df) + 6
        for i, metric in enumerate(df.columns[1:]):  # Skip 'period' column
            chart = LineChart()
            chart.title = metric
            chart.style = 10
            chart.x_axis.title = time_unit.title()
            chart.y_axis.title = metric

            data_ref = Reference(ws, min_col=i + 2, min_row=3, max_row=len(df) + 3)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=len(df) + 3)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 15
            chart.height = 8

            # Position chart
            col_position = ((i % 2) * 9) + 1
            row_position = chart_row + (i // 2) * 16
            ws.add_chart(chart, f"{chr(65 + col_position - 1)}{row_position}")

    def add_category_breakdown_sheet(self, category_col: str, metrics: list,
                                     sheet_name: Optional[str] = None):
        """Add a sheet with category breakdown and bar/pie charts."""
        df = self.processor.aggregate_by_category(category_col, metrics)
        if df.empty:
            return

        sheet_name = sheet_name or f"By {get_friendly_name(category_col)}"
        ws = self.wb.create_sheet(sheet_name)

        # Rename columns
        df = df.rename(columns={col: get_friendly_name(col) for col in df.columns})

        # Add title
        ws.cell(row=1, column=1, value=f"BookSpring Metrics by {get_friendly_name(category_col)}")
        ws.cell(row=1, column=1).font = self.TITLE_FONT

        # Add data
        self._add_dataframe_to_sheet(ws, df, start_row=3)

        # Add bar chart
        chart_row = len(df) + 6
        if len(df) > 0:
            bar_chart = BarChart()
            bar_chart.title = f"Breakdown by {get_friendly_name(category_col)}"
            bar_chart.style = 10

            # Use first numeric column for the bar chart
            data_ref = Reference(ws, min_col=2, min_row=3, max_row=len(df) + 3)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=len(df) + 3)
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.width = 18
            bar_chart.height = 10

            ws.add_chart(bar_chart, f"A{chart_row}")

    def add_comparison_sheet(self, period1_start: date, period1_end: date,
                             period2_start: date, period2_end: date,
                             metrics: list, sheet_name: str = "Period Comparison"):
        """Add a sheet comparing two time periods."""
        df = self.processor.compare_periods(
            period1_start, period1_end, period2_start, period2_end, metrics
        )
        if df.empty:
            return

        ws = self.wb.create_sheet(sheet_name)

        # Rename metric column to friendly names
        df["metric"] = df["metric"].apply(get_friendly_name)

        # Rename columns
        df = df.rename(columns={
            "metric": "Metric",
            "period_1": f"{period1_start} to {period1_end}",
            "period_2": f"{period2_start} to {period2_end}",
            "change": "Change",
            "percent_change": "% Change"
        })

        # Add title
        ws.cell(row=1, column=1, value="Period Comparison")
        ws.cell(row=1, column=1).font = self.TITLE_FONT

        # Add data
        self._add_dataframe_to_sheet(ws, df, start_row=3)

        # Add bar chart comparing periods
        if len(df) > 0:
            chart = BarChart()
            chart.title = "Period Comparison"
            chart.style = 10
            chart.type = "col"

            data_ref = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=len(df) + 3)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=len(df) + 3)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 18
            chart.height = 10

            ws.add_chart(chart, f"A{len(df) + 6}")

    def add_summary_sheet(self):
        """Add a summary statistics sheet."""
        stats = self.processor.get_summary_stats()
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws.cell(row=1, column=1, value="BookSpring Data Summary")
        ws.cell(row=1, column=1).font = self.TITLE_FONT

        # Basic stats
        ws.cell(row=3, column=1, value="Total Records:")
        ws.cell(row=3, column=2, value=stats["total_records"])

        ws.cell(row=4, column=1, value="Date Range:")
        date_range = stats.get("date_range", {})
        start = date_range.get("start", "N/A")
        end = date_range.get("end", "N/A")
        ws.cell(row=4, column=2, value=f"{start} to {end}")

        # Totals section
        ws.cell(row=6, column=1, value="Totals")
        ws.cell(row=6, column=1).font = Font(bold=True)

        row = 7
        for metric, value in stats.get("totals", {}).items():
            ws.cell(row=row, column=1, value=get_friendly_name(metric))
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def save(self, filepath: str):
        """Save the workbook to a file."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(filepath)
        return filepath


def generate_standard_report(processor: DataProcessor, output_path: str,
                             time_unit: TimeUnit = "month") -> str:
    """Generate a standard BookSpring report with common metrics."""
    metrics = [
        "_of_books_distributed",
        "total_children",
        "children_035_months",
        "children_35_years",
        "children_68_years",
        "children_912_years",
        "teens",
        "parents_or_caregivers",
        "minutes_of_activity",
    ]

    # Filter to only metrics that exist in the data
    available_metrics = [m for m in metrics if m in processor.df.columns]

    generator = ExcelReportGenerator(processor)

    # Add summary
    generator.add_summary_sheet()

    # Add time series analysis
    generator.add_time_series_sheet(time_unit, available_metrics)

    # Add breakdowns by category
    if "program" in processor.df.columns:
        generator.add_category_breakdown_sheet("program", available_metrics[:3], "By Program")

    if "activity_type" in processor.df.columns:
        generator.add_category_breakdown_sheet("activity_type", available_metrics[:3], "By Activity Type")

    if "county_served_this_activity" in processor.df.columns:
        generator.add_category_breakdown_sheet("county_served_this_activity",
                                               available_metrics[:2], "By County")

    return generator.save(output_path)
